#CEC300 Navigation Project  
# Charles Gilmore
# Feb. 19, 2022
#
# Project #2 - Quad Copter Navigation
#
# A quadcopter is located at latitude 29.401325 and longitude -81.177222 the quadrotor is oriented so that the y axis is aligned with north/south and the x axis is aligned with east/west. 
# Positive accelerations on the y axis are towards the north and positive accelerations on the x axis are towards the east. Time is in seconds and the acceleration data units are cm/sec^2.
#
# The data is contained in the file quad_accel.xlsx  When the data starts, the quadrotor x velocity is -0.2 m/s and the y velocity is 0.44 m/s 
# The data was taken every 100ms. The accelerometer data is in the file quad_accel_data.xlsx.
#
# Using the acceleration data you are determine the position of the quadcopter at every measurement, convert the position to GPS coordinates, and then plot the flight path on a map of the earth so that the location of the flight is clearly visible. 

#####################################################################

from array import *
from math import *
import random
from turtle import pd
from matplotlib.pyplot import xlabel
import openpyxl
from kivy.garden.mapview import MapView, MapMarker
from kivy.app import App
import csv

## Global Defs
# Initial Conditions
latitude = 29.401325   # (-) West -> (+) East X axis
longitude = -81.177222 # (-) South -> (+) North Y axis
R = 6378137 # Earth's Radius (meters)
init_x_vel = -0.2 # m/s
init_y_vel = 0.44 # m/s
Xaccel_m = 0
Yaccel_m = 0

# Delcare location(path) to excel file > Workbook > Specific Sheet in WB
path = ("quad_accel.xlsx")
wb = openpyxl.load_workbook(path)
sheet = wb.active

# Find max rows & columns of file
max_col = sheet.max_column
max_row = sheet.max_row

# Calculate Long/Lat when given X/Y Displacement
# input: Y Displacement, Init_Long
def findLong(longitude, Ydisp_m):
    # Disp in radians
    long_rad = Ydisp_m/(R * cos((pi * latitude)/180))
    # Offset in Longitude 
    longitude = longitude + long_rad * (180 / pi)
    return longitude

# Calculate Long/Lat when given X/Y Displacement
# input: X Displacement, Init_Lat
def findLat(latitude, Xdisp_m):
    # Disp in radians
    lat_rad = Xdisp_m / R
    latitude = latitude + lat_rad * (180 / pi)
    return latitude

# CSV file creation
fields = ['latitude', 'longitude']
filename = "latlong.csv"
# writing to a csv file
with open(filename, 'w') as csvfile:        
    # creating a csv writer object
    csvwriter = csv.writer(csvfile)
    # writing the fields 
    csvwriter.writerow(fields)
    time=0 
    for i in range(1, max_row):           
        Xdisp_m = (init_x_vel * time) + (0.5 * (sheet.cell(row=i+1, column=2).value * 0.01) * (time ** 2))
        Xlat = findLat(latitude, Xdisp_m)
        Ydisp_m = (init_y_vel * time) + (0.5 * (sheet.cell(row=i+1, column=3).value * 0.01) * (time ** 2))
        Ylong = findLong(longitude, Ydisp_m)
        time+=0.1
        row = [Xlat,Ylong]
        # writing the data fields 
        csvwriter.writerow(row)
         

class MapViewApp(App):
    def build(self):
        mapview = MapView(zoom=11, lat=29.401325, lon=-81.177222)
        # for-in loop to read through the excel data sheet and make calculations
        # Read cell value -> convert cm to m -> calculate displacement from acceleration -> put into array
        time=0
        for i in range(1, max_row):      
            # Convert X acceleration (cm/s^2 > m/s^2) to displacement and put in 1D array
            # Displacement from acceleration formula
            # displ = (Vi)(t) + 1/2(accel)(t^2)
            Xdisp_m = (init_x_vel * time) + (0.5 * (sheet.cell(row=i+1, column=2).value * 0.01) * (time ** 2))
            Xlat = findLat(latitude, Xdisp_m)
    
            # Convert Y acceleration (cm/s^2 > m/s^2) to displacement and put in 1D array
            Ydisp_m = (init_y_vel * time) + (0.5 * (sheet.cell(row=i+1, column=3).value * 0.01) * (time ** 2))
            Ylong = findLong(longitude, Ydisp_m)
    
            time+=0.1
            marker = MapMarker(lat=Xlat, lon=Ylong, source='marker.png')
            mapview.add_marker(marker)
        return mapview
MapViewApp().run()
    